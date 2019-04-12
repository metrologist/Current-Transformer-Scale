from __future__ import division
"""
Read partially processed CT data from Excel to allow Python scripts to tidy
up burden dependence and uncertainty calculations. This is not a replacement
for the Excel CT  templates.
"""

import math
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import warnings
import GTC as gtc

class CALCULATOR(object):
    """
    :param source: is the spreadsheet with data
    :param output: is the name of the spreadsheet that results are placed in.
    """

    def __init__(self, source, output):
        self.source = source
        self.output = output
    
    def makeworkbook(self, set, this_title):
        """
        :param set: a list of lists that are each an excel row
        :param this_title: name of sheet in workbook
        :return: saves the workbook file
        """

        wb = Workbook() #create workbook in memory
        ws0 = wb.active #default active worksheet in position '0'
        ws0.title = this_title
        no_rows = len(set)
        no_columns = len(set[0])
        for i in range(no_rows):
            for j in range(no_columns):
                ws0.cell(row = i+2, column = j+1, value =set[i][j])
        wb.save(filename = self.output)
    
    def getdata_block(self, datasheet,block_range):
        """
        :param datasheet: name of the calc worksheet
        :param block_range: a 4 element list [start row, finish row, start column, finish column]
        :return: a list of lists, each list being the contents of a row
        """

        warnings.simplefilter('ignore')#not interested in 'Discarded range with reserved name'
        wb2 = load_workbook(self.source, data_only = True)#reads numbers rather than formulae
        warnings.simplefilter('default')#turn warnings back on
        sheet = wb2[datasheet]
        selected_rows = []
        for i in range(block_range[0],block_range[1]):
            this_row = []
            for j in range(block_range[2], block_range[3]):
                this_row.append(sheet.cell(row = i, column = j).value)
            selected_rows.append(this_row)
        return selected_rows
        
    def extract_column(self, block, column_number, start_finish):
        """
        For extracting specific table columns after having read an entire sheet with get_datablock

        :param block: in format of self.getdata_block
        :param column_number: Row and column numbers follow R1C1=='A1'
        :param start_finish: list [start row number, finish row number]
        :return: colunmn as a list
        """

        for_output = []
        for i in range(start_finish[0]-1, start_finish[1]): #the -1 includes the first row
            for_output.append(block[i][column_number-1])#-1 needed as first column is 0
        return for_output

    def blockcolumns(self, block):
        """
        converts from imported excel data being lists of rows to being lists of columns

        :param block: list of lists of rows
        :return: list of list of columns
        """

        p = len(block)  # number of rows
        q = len(block[0])  # number of columns
        master = []
        for i in range(q):
            column = []
            for j in range(p):
                column.append(copy_data[j][i])
            master.append(column)
        return master

    def ucomplex_table(self, labels, complex):
        """
        Takes a list of GTC ucomplex results and creates a list of rows where each row has the label, real value,
        expanded uncertainty, coverage factor, imaginary value, expanded uncertainty, coverage factor.

        :param labels: Typically a list of nominal conditions such as the excitation level of a CT
        :param complex: List of GTC ucomplex results matching the list
        :return: a list of rows carrying the values as in the descriptor above
        """

        assert len(labels) == len(complex), "number of labels must match number of ucomplex results"
        table = []
        for i in range(len(labels)):
            row = []
            row.append(labels[i])
            a = complex[i].real
            b = complex[i].imag
            row.append(a.x)  # the real part
            row.append(b.x)  # imaginary value
            ka = gtc.reporting.k_factor(a.df)  # get coverage from dof
            row.append(a.u * ka)  # expanded uncertainty
            kb = gtc.reporting.k_factor(b.df)  # coverage factor from dof
            row.append(b.u * kb)  # expanded uncertainty
            row.append(ka)  # coverage factor
            row.append(kb)  # coverage factot
            table.append(row)

        return table

    def budget_table(self, uncertains, fract):
        """
        For creating a spreadsheet friendly block that presents the uncertainty budget for a list of uncertain numbers

        :param uncertains: a list of uncertain numbers
        :param fract: set to 0 for full budget, but typically 0.1 as fraction of total for inclusion in budget
        :return: a list of row contents, with each row itself a list, suitable for the makeworkbook method
        """
        output_block = []
        for i in range(len(uncertains)):
            header_row = []
            header_row.append('Name')
            header_row.append(uncertains[i].label)
            output_block.append(header_row)
            next_row = []
            next_row.append('Value')
            next_row.append(repr(uncertains[i]))
            output_block.append(next_row)
            next_row = []
            next_row.append('Name')
            next_row.append('Uncertainty contribution')
            output_block.append(next_row)
            for l, u in gtc.rp.budget(uncertains[i], trim = fract ):
                next_row = []
                next_row.append(l)
                next_row.append(u)
                output_block.append(next_row)
            output_block.append(['',''])
            self.makeworkbook(output_block, 'uncertainties')
        return output_block


if __name__ == "__main__":
    ianz = CALCULATOR("S21982 1 turn V5.0_draft_KJ.xlsm", "myResults.xlsx")
    #copy a block from the source worksheet
    block_descriptor = [1,365,1,42]
    my_copy_data = ianz.getdata_block('5 A calc', block_descriptor)
    #put the block, unchanged into the output spreadsheet
    ianz.makeworkbook(my_copy_data,'my_sheet_name')    
    
    
    
    
    